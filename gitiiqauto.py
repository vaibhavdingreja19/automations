import os
import sys
import requests
from typing import Dict, List, Optional
from openpyxl import Workbook

# ---------------- CONFIG ----------------
ORG_NAME = "JHDevOps"                    # GitHub org name
TOKEN_ENV_VAR = "GITHUB_TOKEN"          # Put your PAT in this env var
OUTPUT_XLSX = "github_acl_access.xlsx"  # Output file name
# ----------------------------------------


# How we rank permissions to find "highest"
PERMISSION_RANK = {
    "read": 1,
    "triage": 2,
    "write": 3,
    "maintain": 4,
    "admin": 5,
}

RANK_TO_NAME = {
    1: "Read",
    2: "Triage",
    3: "Write",
    4: "Maintain",
    5: "Admin",
}

API_ROOT = "https://api.github.com"


def get_token() -> str:
    token = os.environ.get(TOKEN_ENV_VAR)
    if not token:
        print(f"Error: GitHub PAT not found in env var {TOKEN_ENV_VAR}.", file=sys.stderr)
        print(f"Set it like:\n  export {TOKEN_ENV_VAR}=your_personal_access_token", file=sys.stderr)
        sys.exit(1)
    return token


def gh_headers(token: str) -> Dict[str, str]:
    return {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github+json",
    }


def paginated_get(url: str, headers: Dict[str, str], params: Dict = None) -> List[Dict]:
    """Simple GitHub API pagination helper."""
    if params is None:
        params = {}
    params = params.copy()
    params.setdefault("per_page", 100)

    results = []
    while url:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code != 200:
            print(f"GitHub API error {resp.status_code} for {url}: {resp.text}", file=sys.stderr)
            break

        results.extend(resp.json())

        # Pagination via Link header
        link = resp.headers.get("Link", "")
        next_url = None
        if link:
            parts = link.split(",")
            for part in parts:
                if 'rel="next"' in part:
                    next_url = part[part.find("<") + 1 : part.find(">")]
                    break

        url = next_url
        params = {}  # only needed on first request

    return results


def get_all_teams(org: str, token: str) -> List[Dict]:
    url = f"{API_ROOT}/orgs/{org}/teams"
    return paginated_get(url, gh_headers(token))


def get_team_repos(org: str, team_slug: str, token: str) -> List[Dict]:
    url = f"{API_ROOT}/orgs/{org}/teams/{team_slug}/repos"
    return paginated_get(url, gh_headers(token))


def highest_permission_for_team(org: str, team_slug: str, token: str) -> Optional[str]:
    repos = get_team_repos(org, team_slug, token)
    if not repos:
        return None

    max_rank = 0

    for repo in repos:
        # Newer GitHub APIs sometimes give "role_name"
        role_name = repo.get("role_name")
        if role_name:
            perm_key = role_name.lower()
        else:
            # Fallback to the legacy boolean permissions object
            perms = repo.get("permissions", {})
            if perms.get("admin"):
                perm_key = "admin"
            elif perms.get("maintain"):
                perm_key = "maintain"
            elif perms.get("push"):
                perm_key = "write"
            elif perms.get("triage"):
                perm_key = "triage"
            elif perms.get("pull"):
                perm_key = "read"
            else:
                # No recognizable permission, skip
                continue

        rank = PERMISSION_RANK.get(perm_key, 0)
        if rank > max_rank:
            max_rank = rank

    if max_rank == 0:
        return None
    return RANK_TO_NAME[max_rank]


def create_excel(teams: List[Dict], token: str, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GitHub ACLs"

    # Header row (A–D)
    ws["A1"] = "Domain"
    ws["B1"] = "GroupName"
    ws["C1"] = "Description"
    ws["D1"] = "AccessLevel"

    row = 2
    for team in teams:
        team_name = team.get("name")      # Human-friendly name
        team_slug = team.get("slug")      # Slug for API calls

        # Always MFCGD in col A
        ws.cell(row=row, column=1, value="MFCGD")
        ws.cell(row=row, column=2, value=team_name)

        # Column C left empty intentionally
        # ws.cell(row=row, column=3).value = ""  # not strictly needed

        highest_perm = highest_permission_for_team(ORG_NAME, team_slug, token)
        if highest_perm:
            ws.cell(row=row, column=4, value=highest_perm)
        else:
            ws.cell(row=row, column=4, value="")  # no repos / no perms

        print(f"{team_name}: {highest_perm}")
        row += 1

    wb.save(output_path)
    print(f"\nExcel written to: {output_path}")


def main():
    token = get_token()
    print(f"Fetching teams for org: {ORG_NAME} ...")
    teams = get_all_teams(ORG_NAME, token)
    print(f"Found {len(teams)} teams.\n")

    create_excel(teams, token, OUTPUT_XLSX)


if __name__ == "__main__":
    main()
