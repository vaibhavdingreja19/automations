import os
import jwt
import time
import requests
from openpyxl import Workbook

BASE_URL = "https://api.github.com"


APP_IDS = [os.getenv('appid1'), os.getenv('appid2'), os.getenv('appid3'), os.getenv('appid4'), os.getenv('appid5'), os.getenv('appid6'), os.getenv('appid7'), os.getenv('appid8')]  
RSA_KEYS = [os.getenv('rsakey1'), os.getenv('rsakey2'), os.getenv('rsakey3'), os.getenv('rsakey4'), os.getenv('rsakey5'), os.getenv('rsakey6'), os.getenv('rsakey7'), os.getenv('rsakey8') ]  


def generate_jwt(app_id, private_key):
    payload = {
        "iat": int(time.time()),
        "exp": int(time.time()) + (10 * 60),
        "iss": app_id,
    }
    token = jwt.encode(payload, private_key, algorithm="RS256")
    return token


def get_installation_token(installation_id, jwt_token):
    url = f"{BASE_URL}/app/installations/{installation_id}/access_tokens"
    headers = {"Authorization": f"Bearer {jwt_token}", "Accept": "application/vnd.github+json"}
    response = requests.post(url, headers=headers)
    response.raise_for_status()
    return response.json()["token"]


def get_installations(jwt_token):
    url = f"{BASE_URL}/app/installations"
    headers = {"Authorization": f"Bearer {jwt_token}", "Accept": "application/vnd.github+json"}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()


def get_repositories(installation_token):
    url = f"{BASE_URL}/installation/repositories"
    headers = {"Authorization": f"Bearer {installation_token}", "Accept": "application/vnd.github+json"}
    all_repositories = []
    page = 1

    while True:
        response = requests.get(url, headers=headers, params={"per_page": 100, "page": page})
        response.raise_for_status()
        repositories = response.json()["repositories"]
        if not repositories:
            break
        all_repositories.extend(repositories)
        page += 1

    return all_repositories


def save_to_excel(data):
    wb = Workbook()
    for app_name, repos in data.items():
        ws = wb.create_sheet(title=app_name[:30]) 
        ws.append(["Repository Name"])
        for repo in repos:
            ws.append([repo])
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save("GitHub_Apps_Repositories.xlsx")


def main():
    if len(APP_IDS) != len(RSA_KEYS):
        raise ValueError("The number of App IDs and RSA private keys must be the same.")

    data = {}
    #for app_id, private_key in zip(APP_IDS, RSA_KEYS):
    for idx, (app_id, private_key) in enumerate(zip(APP_IDS, RSA_KEYS), start=1):
        jwt_token = generate_jwt(app_id, private_key)
        installations = get_installations(jwt_token)

        for installation in installations:
            app_name = installation.get("account", {}).get("login")
            installation_id = installation["id"]

            print(f"Fetching repositories for app: {app_name}")
            installation_token = get_installation_token(installation_id, jwt_token)
            repos = get_repositories(installation_token)
            repo_names = [repo["full_name"] for repo in repos]
            data[f"{app_name}_App{idx}"] = repo_names

    print("Saving data to Excel...")
    save_to_excel(data)
    print("Data saved to 'GitHub_Apps_Repositories.xlsx'")


if __name__ == "__main__":
    main()
