import requests
import openpyxl
from openpyxl.utils import get_column_letter

# === CONFIG ===
GITHUB_TOKEN = "your_token_here"
ORG_NAME = "JHDevOps"
OUTPUT_FILE = "github_repos_branches.xlsx"

headers = {
    "Authorization": f"token {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json"
}

def get_all_repos():
    repos = []
    page = 1
    per_page = 100
    while True:
        url = f"https://api.github.com/orgs/{ORG_NAME}/repos?per_page={per_page}&page={page}&type=all"
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Failed to fetch repos: {response.status_code}, {response.text}")
        data = response.json()
        if not data:
            break
        repos.extend(data)
        page += 1
    return repos

def get_branches(repo_name):
    url = f"https://api.github.com/repos/{ORG_NAME}/{repo_name}/branches"
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        return []
    return response.json()

def save_to_excel(repo_branch_map):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Repos & Branches"
    ws.append(["Repo Name", "Branch Name", "Is Default Branch"])

    for repo, (branches, default_branch) in repo_branch_map.items():
        first = True
        for branch in branches:
            is_default = "default" if branch["name"] == default_branch else ""
            if first:
                ws.append([repo, branch["name"], is_default])
                first = False
            else:
                ws.append(["", branch["name"], is_default])

    # Auto-width columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    wb.save(OUTPUT_FILE)
    print(f"Saved to {OUTPUT_FILE}")

def main():
    print("Fetching repositories...")
    repos = get_all_repos()
    print(f"Total repos found: {len(repos)}")

    repo_branch_map = {}
    for repo in repos:
        name = repo["name"]
        default_branch = repo.get("default_branch", "")
        print(f"Fetching branches for repo: {name}")
        branches = get_branches(name)
        repo_branch_map[name] = (branches, default_branch)

    print("Saving to Excel...")
    save_to_excel(repo_branch_map)
    print("Done.")

if __name__ == "__main__":
    main()
